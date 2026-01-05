import openpyxl
import pandas as pd
from typing import Dict, Any, List
from app.parsers.base_parser import BaseParser

class ExcelParser(BaseParser):
    """Parser for Excel (.xlsx) documents"""
    
    def parse(self) -> Dict[str, Any]:
        """Parse Excel file and extract structured data"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            self.extracted_data = {
                'file_type': 'Excel',
                'sheets': list(wb.sheetnames),
                'data': self._extract_all_sheets(wb),
                'sheet_details': self._extract_sheet_details(wb)
            }
            wb.close()
            return self.extracted_data
        except Exception as e:
            raise Exception(f"Error parsing Excel file: {str(e)}")
    
    def extract_text(self) -> str:
        """Extract all text from Excel"""
        try:
            df_dict = pd.read_excel(self.file_path, sheet_name=None)
            text = ""
            for sheet_name, df in df_dict.items():
                text += f"\n--- SHEET: {sheet_name} ---\n"
                text += df.to_string() + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error extracting text from Excel: {str(e)}")
    
    def _extract_all_sheets(self, wb) -> Dict[str, List[List[Any]]]:
        """Extract data from all sheets"""
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []
            for row in ws.iter_rows(values_only=True):
                sheet_data.append(list(row))
            sheets_data[sheet_name] = sheet_data
        return sheets_data
    
    def _extract_sheet_details(self, wb) -> Dict[str, Dict[str, Any]]:
        """Extract detailed information about each sheet"""
        details = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            details[sheet_name] = {
                'dimensions': ws.dimensions,
                'max_row': ws.max_row,
                'max_column': ws.max_column
            }
        return details
